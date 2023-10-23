from requests_html import HTMLSession
import requests
import json
import urllib
import openpyxl
import pandas as pd
import time
from openpyxl import load_workbook
from datetime import datetime
import os.path   
import random
import pprint
import datetime
from collections import Counter
import numpy as np
import pprint

kbl_mkdir_path = 'D:/pythonAppCode/ProtoAnalysis/KBO/'

TEAM_ID = {'부산KT':'06',
           '울산모비':'10',
           '원주DB':'16',
           '서울삼성':'35',
           '창원LG':'50',
           '서울SK':'55',
           '전주KCC':'60',
           '안양KGC':'70',
           '한국가스':'64',
           '고양캐롯':'73'}
             

# KBL 공홈 팀별 데이터 크롤링 2019 ~ 2023
def kbl_Team_DataLoad(year):

    fileName = kbl_mkdir_path + 'KBL_팀_'+ str(year) +'시즌별_통합_데이터.xlsx'

    print(fileName)

    wb = openpyxl.Workbook()
    ws = wb.active
    wb.save(fileName)  

    if year == 2023: year = 41
    elif year == 2022: year = 39
    elif year == 2021: year = 37
    elif year == 2020: year = 35

    url = f"https://api.kbl.or.kr/leagues/S{str(year)}G01/stats/teams?"

    headers = {
        'Origin':'https://www.kbl.or.kr',
        'Host': 'api.kbl.or.kr',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36',
        'X-Requested-With': 'XMLHttpRequest',
        'sec-ch-ua': '"Chromium";v="110", "Not A(Brand";v="24", "Google Chrome";v="110"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': "Windows",
        'Sec-Fetch-Dest': 'empty',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Site': 'same-site',
        'TeamCode': 'XX',
        'Accept': 'application/json, text/plain, */*',
        'Accept-Encoding': 'gzip, deflate, br',
        'Accept-Language': 'ko,ko-KR;q=0.9,en-US;q=0.8,en;q=0.7,ja;q=0.6',
        'Access-Control-Allow-Credentials': 'true',
        'Access-Control-Allow-Methods': 'GET, PUT, POST, DELETE, OPTIONS',
        'Access-Control-Allow-Origin': '*',
        'Cache-Control': 'no-cache',
        'Channel': 'WEB',
        'Connection': 'keep-alive',
        'Pragma': 'no-cache',
    }
    
    with HTMLSession() as s:
        result = s.get(url, headers=headers).text
        json_result = json.loads(result)

        teamIndex = 1
        for output in json_result:

            if teamIndex == 1:
                teamInsideIndex = 8
                ws.append(['팀ID','팀명','팀 풀네임','경기수','승','무','패'])
                for category in output['records'].keys():
                    ws.cell(teamIndex, teamInsideIndex).value = category
                    teamInsideIndex += 1
                teamIndex += 1
                
            tcode = output['team']['tcode']
            tname = output['team']['tname']
            tnameF = output['team']['tnameF']

            gameCount = output['gameCount']
            win = output['win']
            draw = output['draw']
            loss = output['loss']

            ws.cell(teamIndex, 1).value = tcode
            ws.cell(teamIndex, 2).value = tname
            ws.cell(teamIndex, 3).value = tnameF
            ws.cell(teamIndex, 4).value = gameCount
            ws.cell(teamIndex, 5).value = win
            ws.cell(teamIndex, 6).value = draw
            ws.cell(teamIndex, 7).value = loss

            teamInsideIndex = 8
            for records in output['records'].values():
                ws.cell(teamIndex, teamInsideIndex).value = records
                teamInsideIndex += 1

            wb.save(fileName)
            teamIndex += 1

# KBL 공홈 선수별 데이터 크롤링 2019 ~ 2023
def kbl_Player_DataLoad(year):

    fileName = kbl_mkdir_path + 'KBL_선수_'+ str(year) +'시즌별_통합_데이터.xlsx'

    print(fileName)

    wb = openpyxl.Workbook()
    ws = wb.active
    wb.save(fileName)  

    if year == 2023: year = 41
    elif year == 2022: year = 39
    elif year == 2021: year = 37
    elif year == 2020: year = 35

    url = f"https://api.kbl.or.kr/leagues/S{str(year)}G01/stats/players?tcodeList=all"

    headers = {
        'Origin':'https://www.kbl.or.kr',
        'Host': 'api.kbl.or.kr',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36',
        'X-Requested-With': 'XMLHttpRequest',
        'sec-ch-ua': '"Chromium";v="110", "Not A(Brand";v="24", "Google Chrome";v="110"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': "Windows",
        'Sec-Fetch-Dest': 'empty',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Site': 'same-site',
        'TeamCode': 'XX',
        'Accept': 'application/json, text/plain, */*',
        'Accept-Encoding': 'gzip, deflate, br',
        'Accept-Language': 'ko,ko-KR;q=0.9,en-US;q=0.8,en;q=0.7,ja;q=0.6',
        'Access-Control-Allow-Credentials': 'true',
        'Access-Control-Allow-Methods': 'GET, PUT, POST, DELETE, OPTIONS',
        'Access-Control-Allow-Origin': '*',
        'Cache-Control': 'no-cache',
        'Channel': 'WEB',
        'Connection': 'keep-alive',
        'Pragma': 'no-cache',
    }
    
    with HTMLSession() as s:
        result = s.get(url, headers=headers).text
        json_result = json.loads(result)

        teamIndex = 1
        for output in json_result:

            if teamIndex == 1:
                teamInsideIndex = 6
                ws.append(['플레이어 ID','플레이어명','팀 ID','팀 명','경기수'])
                for category in output['records'].keys():
                    ws.cell(teamIndex, teamInsideIndex).value = category
                    teamInsideIndex += 1
                teamIndex += 1
                
            pcode = output['player']['pcode']
            pname = output['player']['pname']
            tcode = output['player']['tcode']
            tname = output['player']['tname']

            gameCount = output['gameCount']

            if gameCount == 0: continue

            ws.cell(teamIndex, 1).value = pcode
            ws.cell(teamIndex, 2).value = pname
            ws.cell(teamIndex, 3).value = tcode
            ws.cell(teamIndex, 4).value = tname
            ws.cell(teamIndex, 5).value = gameCount

            teamInsideIndex = 6
            for records in output['records'].values():
                ws.cell(teamIndex, teamInsideIndex).value = records
                teamInsideIndex += 1

            wb.save(fileName)
            teamIndex += 1

# KBL 팀 스탯 통계
def kbl_Team_Analysis():

    yearList = ['2020','2021','2022','2023']

    save_fileName = kbl_mkdir_path + 'KBL_팀_스탯_분석_데이터.xlsx'
    print(save_fileName)

    save_wb = openpyxl.Workbook()
    save_ws = save_wb.active
    save_ws.append(['팀ID','팀 명','2019-20시즌','2020-21시즌','2021-22시즌','2022-23시즌', '총합'])
    save_wb.save(save_fileName) 

    for year in yearList:

        loadfileName = kbl_mkdir_path + 'KBL_팀_' + year + '시즌별_통합_데이터.xlsx'

        load_wb = load_workbook(filename = loadfileName, data_only=True)
        ws = load_wb[load_wb.sheetnames[0]]

        print(loadfileName)

        book = openpyxl.load_workbook(save_fileName)
        sheet = book.worksheets[0]

        index = 1
        for row in ws.rows:

            if index == 1:
                index += 1
                continue

            _teamID = row[0].value # 팀 ID
            _teamName = row[1].value.replace(' ','') # 팀명

            _pts = row[32].value # 득점
            _oreb = row[25].value # 공격 리바운드
            _dreb = row[9].value # 수비 리바운드
            _gd = row[21].value # 굿 디펜스
            _reb = row[31].value # 총 리바운드
            _fgm = row[14].value # 야투 성공 횟수
            _fga = row[17].value # 야투 시도 횟수
            _fta = row[20].value # 자유투 시도 횟수
            _ftm = row[19].value # 자유투 성공 횟수
            _stl = row[33].value # 스틸
            _ast = row[7].value # 어시스트
            _blk = row[8].value # 블록
            _tov = row[37].value # 실책(턴오버)
            _pf = row[18].value # 파울

            # KBL 팀 스코어
            _kblTeamScore = _pts + 0.4*_fgm - 0.7*_fga - 0.4*(_fta - _ftm) + 0.7*_oreb + 0.3*_dreb + _stl +0.7*_ast + 0.7*_blk - 0.4*_pf - _tov

            if year == '2020': 
                save_ws.cell(index, 1).value = _teamID
                save_ws.cell(index, 2).value = _teamName
                save_ws.cell(index, 3).value = _kblTeamScore
            elif year == '2021':
                saveNumber = 1
                checkSave = False
                for saveRow in sheet.rows:
                    if saveRow[0].value == _teamID:
                        save_ws.cell(saveNumber, 4).value = _kblTeamScore
                        checkSave = True
                        break
                    saveNumber += 1

                if checkSave == False:
                    save_ws.append([_teamID,_teamName,0,_kblTeamScore])
            elif year == '2022':
                saveNumber = 1
                checkSave = False
                for saveRow in sheet.rows:
                    if saveRow[0].value == _teamID:
                        save_ws.cell(saveNumber, 5).value = _kblTeamScore
                        checkSave = True
                        break
                    saveNumber += 1

                if checkSave == False:
                    save_ws.append([_teamID,_teamName,0,0,_kblTeamScore])
            elif year == '2023':
                saveNumber = 1
                checkSave = False
                for saveRow in sheet.rows:
                    if saveRow[0].value == _teamID:
                        save_ws.cell(saveNumber, 6).value = _kblTeamScore
                        checkSave = True
                        break
                    saveNumber += 1
                if checkSave == False:
                    save_ws.append([_teamID,_teamName,0,0,0,_kblTeamScore])

            index += 1
            save_wb.save(save_fileName)  

    #엑셀파일 불러오기
    wb = load_workbook(save_fileName)
    ws = wb.active

    index = 1
    for row in ws.rows:
        if index == 1:
            index += 1
            continue

        dataInCheck = 4

        season20 = row[2].value
        if season20 == None: 
            season20 = 0
            dataInCheck -= 1
        season21 = row[3].value
        if season21 == None: 
            season21 = 0
            dataInCheck -= 1
        season22 = row[4].value
        if season22 == None: 
            season22 = 0
            dataInCheck -= 1
        season23 = row[5].value
        if season23 == None: 
            season23 = 0
            dataInCheck -= 1

        # 19~20 - 0.1 / 20~21 - 0.2 / 21~22 - 0.3 / 22~23 - 0.4
        result = ((season20*0.1) + (season21*0.2) + (season22*0.3) + (season23*4)) / dataInCheck

        ws.cell(index, 7).value = result

        index += 1
    
    wb.save(save_fileName)

# KBL 선수 스탯 통계
def kbl_Player_Analysis():

    yearList = ['2020','2021','2022','2023']

    save_fileName = kbl_mkdir_path + 'KBL_선수_스탯_분석_데이터.xlsx'
    print(save_fileName)

    save_wb = openpyxl.Workbook()
    save_ws = save_wb.active
    save_ws.append(['플레이어ID','플레이어 명','팀ID','팀 명','2019-20시즌','2020-21시즌','2021-22시즌','2022-23시즌', '총합'])
    save_wb.save(save_fileName) 

    for year in yearList:

        loadfileName = kbl_mkdir_path + 'KBL_선수_' + year + '시즌별_통합_데이터.xlsx'

        load_wb = load_workbook(filename = loadfileName, data_only=True)
        ws = load_wb[load_wb.sheetnames[0]]

        print(loadfileName)

        book = openpyxl.load_workbook(save_fileName)
        sheet = book.worksheets[0]

        index = 1
        for row in ws.rows:

            if index == 1:
                index += 1
                continue

            _playerID = row[0].value # 플레이어 ID
            _PlayerName = row[1].value # 플레이어 명
            _teamID = row[2].value # 팀 ID
            _teamName = row[3].value.replace(' ','') # 팀명

            _pts = row[30].value # 득점
            _oreb = row[23].value # 공격 리바운드
            _dreb = row[7].value # 수비 리바운드
            _gd = row[19].value # 굿 디펜스
            _reb = row[29].value # 총 리바운드
            _fgm = row[12].value # 야투 성공 횟수
            _fga = row[15].value # 야투 시도 횟수
            _fta = row[18].value # 자유투 시도 횟수
            _ftm = row[17].value # 자유투 성공 횟수
            _stl = row[31].value # 스틸
            _ast = row[5].value # 어시스트
            _blk = row[6].value # 블록
            _tov = row[35].value # 실책(턴오버)
            _pf = row[16].value # 파울
            _playMin = row[24].value # 출전시간(분)

            # 가산법
            _kblEfficiencyAdd = (_pts+_stl+_blk+_dreb)*1.0 + (_oreb+_ast+_gd)*1.5 + _playMin / 4
            _kblEfficiencyMinus = _tov*1.5 + (_fga-_fgm)*1.0+(_fta-_ftm)*0.8

            # KBL 플레이어 가성비
            _kblPlayerEfficiency = _kblEfficiencyAdd - _kblEfficiencyMinus

            if year == '2020': 
                save_ws.cell(index, 1).value = _playerID
                save_ws.cell(index, 2).value = _PlayerName
                save_ws.cell(index, 3).value = _teamID
                save_ws.cell(index, 4).value = _teamName
                save_ws.cell(index, 5).value = _kblPlayerEfficiency
            elif year == '2021':
                saveNumber = 1
                checkSave = False
                for saveRow in sheet.rows:
                    if saveRow[0].value == _playerID:
                        save_ws.cell(saveNumber, 6).value = _kblPlayerEfficiency
                        checkSave = True
                        break
                    saveNumber += 1

                if checkSave == False:
                    save_ws.append([_playerID,_PlayerName,_teamID,_teamName,0,_kblPlayerEfficiency])
            elif year == '2022':
                saveNumber = 1
                checkSave = False
                for saveRow in sheet.rows:
                    if saveRow[0].value == _playerID:
                        save_ws.cell(saveNumber, 7).value = _kblPlayerEfficiency
                        checkSave = True
                        break
                    saveNumber += 1

                if checkSave == False:
                    save_ws.append([_playerID,_PlayerName,_teamID,_teamName,0,0,_kblPlayerEfficiency])
            elif year == '2023':
                saveNumber = 1
                checkSave = False
                for saveRow in sheet.rows:
                    if saveRow[0].value == _playerID:
                        save_ws.cell(saveNumber, 8).value = _kblPlayerEfficiency
                        checkSave = True
                        break
                    saveNumber += 1
                if checkSave == False:
                    save_ws.append([_playerID,_PlayerName,_teamID,_teamName,0,0,0,_kblPlayerEfficiency])

            index += 1
            save_wb.save(save_fileName)  

    #엑셀파일 불러오기
    wb = load_workbook(save_fileName)
    ws = wb.active

    index = 1
    for row in ws.rows:
        if index == 1:
            index += 1
            continue

        dataInCheck = 4

        season20 = row[4].value
        if season20 == None: 
            season20 = 0
            dataInCheck -= 1
        season21 = row[5].value
        if season21 == None: 
            season21 = 0
            dataInCheck -= 1
        season22 = row[6].value
        if season22 == None: 
            season22 = 0
            dataInCheck -= 1
        season23 = row[7].value
        if season23 == None: 
            season23 = 0
            dataInCheck -= 1

        # 19~20 - 0.1 / 20~21 - 0.2 / 21~22 - 0.3 / 22~23 - 0.4
        result = ((season20*0.1) + (season21*0.2) + (season22*0.3) + (season23*0.4)) / dataInCheck

        ws.cell(index, 9).value = result

        index += 1
    
    wb.save(save_fileName)

# KBL 전반전 점수 통계
def kbl_FirstHalf_Data(homeID, awayID):

    homeScore = 0
    awayScore = 0

    # 2020/2021/2022/2023
    yearList = [35,37,39,41]


    index = 1
    for year in yearList:

        url = f'https://api.kbl.or.kr/leagues/S{str(year)}G01/vs/team-stats?teamA={homeID}&teamB={awayID}'

        headers = {
            'Accept': 'application/json, text/plain, */*',
            'Accept-Encoding': 'gzip, deflate, br',
            'Accept-Language': 'ko,ko-KR;q=0.9,en-US;q=0.8,en;q=0.7,ja;q=0.6',
            'Access-Control-Allow-Credentials': 'true',
            'Access-Control-Allow-Methods': 'GET, PUT, POST, DELETE, OPTIONS',
            'Access-Control-Allow-Origin': '*',
            'Cache-Control': 'no-cache',
            'Channel': 'WEB',
            'Connection': 'keep-alive',
            'Host': 'api.kbl.or.kr',
            'Origin': 'https://www.kbl.or.kr',
            'Pragma': 'no-cache',
            'sec-ch-ua': '"Chromium";v="110", "Not A(Brand";v="24", "Google Chrome";v="110"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': 'Windows',
            'Sec-Fetch-Dest': 'empty',
            'Sec-Fetch-Mode': 'cors',
            'Sec-Fetch-Site': 'same-site',
            'TeamCode': 'XX',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36',
            'X-Requested-With': 'XMLHttpRequest',
        }
        
        with HTMLSession() as s:
            result = s.get(url, headers=headers).text
            json_result = json.loads(result)

            try:
                homeGameCount = json_result['a']['gameCount']
            except:
                continue
            awayGameCount = json_result['b']['gameCount']

            homeScoreq1 = json_result['a']['scores']['scoreq1']
            homeScoreq2 = json_result['a']['scores']['scoreq2']

            awayScoreq1 = json_result['b']['scores']['scoreq1']
            awayScoreq2 = json_result['b']['scores']['scoreq2']

            homeScore = homeScore + int((homeScoreq1 / homeGameCount) + (homeScoreq2 / homeGameCount))
            awayScore = awayScore + int((awayScoreq1 / awayGameCount) + (awayScoreq2 / awayGameCount))

            index += 1

        time.sleep(1)

    scoreList = []
    scoreList.append(int(homeScore/index))
    scoreList.append(int(awayScore/index))

    return scoreList

# 배트맨 KBL 2019~2023 상대 전적 (홈, 원정 동일 위치) - typeText 'Home', 'Away'
def betman_vs_Score(typeText, homeID, awayID):

    yearList = [2019,2020,2021,2022,2023]

    teamScoreList = []

    for year in yearList:

        loadfileName = kbl_mkdir_path + '배트맨_KBL_' + str(year) + '년도_데이터.xlsx'

        wb = load_workbook(filename = loadfileName, data_only=True)
        ws = wb[wb.sheetnames[0]]

        for row in ws.rows:

            _date = str(row[2].value)

            _homeID = str(row[4].value)
            _awayID = str(row[5].value)

            _matchResult = str(row[15].value)

            if _homeID != homeID or _awayID != awayID: continue

            if typeText == 'Home':
                teamScoreList.append(_matchResult.split(':')[0])
            elif typeText == 'Away':
                teamScoreList.append(_matchResult.split(':')[1])
                    
    teamScore = 0
    for score in teamScoreList:
        teamScore += int(score)

    return teamScore/len(teamScoreList)

# 배트맨 KBL 2022~2023 최근 전적 (홈, 원정 동일 위치) - typeText 'Home', 'Away'
def betman_RecentScore(typeText, teamID):

    yearList = [2022,2023]

    teamScoreList = []

    for year in yearList:

        loadfileName = kbl_mkdir_path + '배트맨_KBL_' + str(year) + '년도_데이터.xlsx'

        wb = load_workbook(filename = loadfileName, data_only=True)
        ws = wb[wb.sheetnames[0]]

        for row in ws.rows:

            # 0='No', 1='종류', 2='날짜', 3='리그 이름', 4='홈 네임', 5='원정 네임', 6='승리 배당'
            # 7='무 배당', 8='패배 배당', 9='홈 핸디', 10='무 핸디', 11='원정 핸디', 12='경기 결과', 13='경기 결과 점수'

            _date = str(row[2].value)

            _homeID = str(row[4].value)
            _awayID = str(row[5].value)
            _matchResult = str(row[15].value)

            if typeText == 'Home':
                if _homeID == teamID: 
                    teamScoreList.append(_matchResult.split(':')[0])
            elif typeText == 'Away':
                if _awayID == teamID: 
                    teamScoreList.append(_matchResult.split(':')[1])
                    
    teamScore = 0
    for score in teamScoreList:
        teamScore += int(score)

    return teamScore/len(teamScoreList)

# 배트맨 KBL 2019~2023 상대 점수차 (홈, 원정 동일 위치)
def betman_Difference_Score(homeID, awayID):

    yearList = [2019,2020,2021,2022,2023]

    teamScoreList = []

    for year in yearList:

        loadfileName = kbl_mkdir_path + '배트맨_KBL_' + str(year) + '년도_데이터.xlsx'

        wb = load_workbook(filename = loadfileName, data_only=True)
        ws = wb[wb.sheetnames[0]]

        for row in ws.rows:

            # 0='No', 1='종류', 2='날짜', 3='리그 이름', 4='홈 네임', 5='원정 네임', 6='승리 배당'
            # 7='무 배당', 8='패배 배당', 9='홈 핸디', 10='무 핸디', 11='원정 핸디', 12='경기 결과', 13='경기 결과 점수'

            _homeID = str(row[4].value)
            _awayID = str(row[5].value)

            _matchResult = str(row[15].value)

            if _homeID != homeID or _awayID != awayID: continue

            result = int(_matchResult.split(':')[0]) - int(_matchResult.split(':')[1])

            if result < 0: teamScoreList.append(-result)
            else: teamScoreList.append(result)

                    
    teamScore = 0
    for score in teamScoreList:
        teamScore += int(score)

    return teamScore/len(teamScoreList)






# 배트맨 KBO 데이터 크롤링
def betman_KBO_DataLoad(year):

    fileName = kbl_mkdir_path + '배트맨_KBO_' + str(year) + '년도_데이터.xlsx'

    print(fileName)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['No', '종류', '날짜', '리그 이름','홈팀 ID','원정팀 ID','홈 네임', '원정 네임','승리 배당','무 배당','패배 배당','홈 핸디','무 핸디','원정 핸디','경기 결과','경기 결과 점수'])

    wb.save(fileName)  

    year = str(year)[2:]

    url = "https://www.betman.co.kr/buyPsblGame/gameInfoInq.do"

    # 210001 = 21년도 / 0001 -> 1주차 경기
    dataNumbering = int(year+'0001')

    row = 2
    for i in range(200):

        print('데이터 주차 : ' + str(dataNumbering))

        headers = {
            "Referer": f"https://www.betman.co.kr/main/mainPage/gamebuy/gameSlipIFR.do?gmId=G101&gmTs={str(dataNumbering)}&gameDivCd=C&isIFR=Y",            
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36",
            'accept': '*/*',
            'accept-encoding': 'gzip, deflate, br',
            'accept-language': 'ko',
            'Content-Type': 'application/json'
        }

        data = {
            "gmId" : "G101",
            "gmTs" : str(dataNumbering),
            "gameYear" : "",
            "_sbmInfo" : {"_sbmInfo" : {"debugMode": "false"}}
        }

        with HTMLSession() as s:
            result = s.post(url, headers=headers, data=json.dumps(data)).text
            json_result = json.loads(result)

            try:
                output = json_result['compSchedules']['datas']
            except:
                print(str(dataNumbering) + ' : 데이터가 없습니다.')
                break

            dataNumbering += 1

            for li in output:

                _type = li[0] # 스포츠 종류 (SC=축구/BK=농구/VL=배구/BS=야구)

                # 경기 날짜, 시간
                date = str(li[3])[:len(str(li[3]))-3] # 13자리에서 뒤에 3자리 빼고 변환(파이썬에서는 그렇게 해야 변환되서)
                _matchDate = datetime.datetime.fromtimestamp(int(date)).strftime('%Y-%m-%d %H:%M:%S')

                _leagueName = li[7] # 리그 이름
                _homeID = li[12] # 홈팀 ID
                _awayID = li[13] # 원정팀 ID
                _homeName = li[14] # 홈 네임
                _awayName = li[15] # 원정 네임
                _winRate = li[16] # 승리 배당률
                _drawRate = li[17] # 무승부 배당률
                _lossRate = li[18] # 패배 배당률

                _homeHandy = li[20] # 홈 핸디
                _drawHandy = li[21] # 무 핸디
                _awayHandy = li[22] # 원정 핸디

                _result = li[28] # 경기 결과 - 홈 0 / 무 1 / 원정 2

                _matchResult = li[33] # 경기 결과 점수

                if _type != 'BS' or _matchResult == None or _homeHandy != 0.0 or _leagueName != 'KBO': 
                    continue

                ## 엑셀 저장
                ws.cell(row, 1).value = row-1
                ws.cell(row, 2).value = _type
                ws.cell(row, 3).value = _matchDate
                ws.cell(row, 4).value = _leagueName
                ws.cell(row, 5).value = _homeID
                ws.cell(row, 6).value = _awayID
                ws.cell(row, 7).value = _homeName
                ws.cell(row, 8).value = _awayName

                ws.cell(row, 9).value = _winRate
                ws.cell(row, 10).value = _drawRate
                ws.cell(row, 11).value = _lossRate
                ws.cell(row, 12).value = _homeHandy
                ws.cell(row, 13).value = _drawHandy
                ws.cell(row, 14).value = _awayHandy
                ws.cell(row, 15).value = _result
                ws.cell(row, 16).value = _matchResult
                row += 1

    wb.save(fileName)









# 베트맨 금일 라인업 통계
def betman_DailyLineup_Analysis(homeName, awayName):
    
    teamFileName = kbl_mkdir_path + 'KBL_팀_스탯_분석_데이터.xlsx'
    playerFileName = kbl_mkdir_path + 'KBL_선수_스탯_분석_데이터.xlsx'

    player_wb = load_workbook(filename = playerFileName, data_only=True)
    player_ws = player_wb[player_wb.sheetnames[0]]

    team_wb = load_workbook(filename = teamFileName, data_only=True)
    team_ws = team_wb[team_wb.sheetnames[0]]

    print(awayName)

    if homeName == '울산모비': homeName = '울산현대모비스'
    if awayName == '울산모비': awayName = '울산현대모비스'
    
    if homeName == '한국가스': homeName = '대구한국가스공사'
    if awayName == '한국가스': awayName = '대구한국가스공사'

    if homeName == 'KT소닉붐': homeName = '부산KT'
    if awayName == 'KT소닉붐': awayName = '부산KT'

    print(awayName)

    # 플레이어
    homePlayerIndex = 1
    awayPlayerIndex = 1

    homePlayerScore = 0
    awayPlayerScore = 0
    for row in player_ws.rows:
        if row[3].value == homeName and row[7].value != None:
            homePlayerScore += row[8].value
            homePlayerIndex += 1

        if row[3].value == awayName and row[7].value != None:
            awayPlayerScore += row[8].value
            awayPlayerIndex += 1

    # 팀
    homeTeamScore = 1
    awayTeamScore = 1
    for row in team_ws.rows:
        if row[1].value == homeName and row[5].value != None:
            homeTeamScore = int(row[6].value)

        if row[1].value == awayName and row[5].value != None:
            awayTeamScore = int(row[6].value)

        
    # (상대전적x0.65) + (최근전적x0.35)
    vsScore = betman_vs_Score('Home',TEAM_ID[homeName],TEAM_ID[awayName])
    recentScore = betman_RecentScore('Home',TEAM_ID[homeName])

    homeScore = (vsScore*0.65) + (recentScore*0.35)

    # (상대전적x0.65) + (최근전적x0.35)
    vsScore = betman_vs_Score('Away',TEAM_ID[homeName],TEAM_ID[awayName])
    recentScore = betman_RecentScore('Away',TEAM_ID[awayName])

    awayScore = (vsScore*0.65) + (recentScore*0.35)

    homePlayerScore = int((homePlayerScore / homePlayerIndex))
    awayPlayerScore = int((awayPlayerScore / awayPlayerIndex))

    # 홈 팀에 50점 더 추가로 줌
    homeResult = homePlayerScore + 50 + homeTeamScore
    awayResult = awayPlayerScore + awayTeamScore

    gapScore = int(betman_Difference_Score(TEAM_ID[homeName],TEAM_ID[awayName]))

    halfList = kbl_FirstHalf_Data(TEAM_ID[homeName], TEAM_ID[awayName])

    halfGapScore = halfList[0] - halfList[1]

    if halfGapScore < 0: halfGapScore = -halfGapScore

    if homeResult > awayResult:
        changeHomeScore = int(homeScore)
        changeAwayScore = changeHomeScore - gapScore

        changeHomeHalfScore = halfList[0]
        changeAwayHalfScore = halfList[0] - halfGapScore
    else:
        changeAwayScore = int(awayScore)
        changeHomeScore = changeAwayScore - gapScore

        changeAwayHalfScore = halfList[1]
        changeHomeHalfScore = halfList[1] - halfGapScore
        
    print(homeName + ' : ' + awayName)
    print('기본 전반전 점수 -> ' + str(halfList[0]) + ' : ' + str(halfList[1]))
    print('스탯 점수 추가 전반전 -> ' + str(changeHomeHalfScore) + ' : ' + str(changeAwayHalfScore))
    print('기본 -> ' + str(int(homeScore)) + ' : ' + str(int(awayScore)))
    print('스탯 점수 추가 후 -> ' + str(changeHomeScore) + ' : ' + str(changeAwayScore))
    print('총 합 스탯 점수 -> ' + str(homeResult) + ' : ' + str(awayResult))
    print('팀 스탯 점수 -> ' + str(homeTeamScore) + ' : ' + str(awayTeamScore))
    print('플레이어 스탯 점수 -> ' + str(homePlayerScore) + ' : ' + str(awayPlayerScore))
    print('■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■')

if __name__ == '__main__':

    # ■■■■■■■■■■ 매일 업데이트 ■■■■■■■■■■
    
    # kbl_Team_DataLoad(2023)
    # kbl_Player_DataLoad(2023)
    betman_KBO_DataLoad(2019)
    betman_KBO_DataLoad(2020)
    betman_KBO_DataLoad(2021)
    betman_KBO_DataLoad(2022)
    betman_KBO_DataLoad(2023)

    # kbl_Team_Analysis()
    # kbl_Player_Analysis()

    # ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■

    # 배트맨 라인업
    # betman_DailyLineup_Analysis('전주KCC','창원LG')

