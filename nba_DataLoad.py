from requests_html import HTMLSession
import requests
import json
import urllib
import openpyxl
import pandas as pd
import time
from wordpress_xmlrpc import Client
from wordpress_xmlrpc import WordPressPost
from wordpress_xmlrpc.methods import posts
from openpyxl import load_workbook
from datetime import datetime
import os.path   
import random
import pprint
import datetime
from collections import Counter
import numpy as np

nba_mkdir_path = 'D:/pythonAppCode/ProtoAnalysis/nba/'

TEAM_KOR_TO = {'멤피그리':'Memphis Grizzlies',
             '덴버너게':'Denver Nuggets',
             '필라76s':'Philadelphia 76ers',
             '피닉선즈':'Phoenix Suns',
             '보스셀틱':'Boston Celtics',
             '샬럿호네':'Charlotte Hornets',
             '뉴올펠리':'New Orleans Pelicans',
             '토론랩터':'Toronto Raptors',
             '브루네츠':'Brooklyn Nets',
             '애틀호크':'Atlanta Hawks',
             '시카불스':'Chicago Bulls',
             '클리캐벌':'Cleveland Cavaliers',
             '샌안스퍼':'San Antonio Spurs',
             '인디페이':'Indiana Pacers',
             '밀워벅스':'Milwaukee Bucks',
             '골든워리':'Golden State Warriors',
             '워싱위저':'Washington Wizards',
             '마이히트':'Miami Heat',
             '미네울브':'Minnesota Timberwolves',
             'LA레이커':'Los Angeles Lakers',
             'LA클리퍼':'LA Clippers',
             '유타재즈':'Utah Jazz',
             '새크킹스':'Sacramento Kings',
             '오클썬더':'Oklahoma City Thunder',
             '뉴욕닉스':'New York Knicks',
             '올랜매직':'Orlando Magic',
             '포틀트레':'Portland Trail Blazers',
             '댈러매버':'Dallas Mavericks',
             '디트피스':'Detroit Pistons',
             '휴스로케':'Houston Rockets'}

TEAM_EN_TO = {'Memphis Grizzlies':'멤피그리',
             'Denver Nuggets':'덴버너게',
             'Philadelphia 76ers':'필라76s',
             'Phoenix Suns':'피닉선즈',
             'Boston Celtics':'보스셀틱',
             'Charlotte Hornets':'샬럿호네',
             'New Orleans Pelicans':'뉴올펠리',
             'Toronto Raptors':'토론랩터',
             'Brooklyn Nets':'브루네츠',
             'Atlanta Hawks':'애틀호크',
             'Chicago Bulls':'시카불스',
             'Cleveland Cavaliers':'클리캐벌',
             'San Antonio Spurs':'샌안스퍼',
             'Indiana Pacers':'인디페이',
             'Milwaukee Bucks':'밀워벅스',
             'Golden State Warriors':'골든워리',
             'Washington Wizards':'워싱위저',
             'Miami Heat':'마이히트',
             'Minnesota Timberwolves':'미네울브',
             'Los Angeles Lakers':'LA레이커',
             'LA Clippers':'LA클리퍼',
             'Utah Jazz':'유타재즈',
             'Sacramento Kings':'새크킹스',
             'Oklahoma City Thunder':'오클썬더',
             'New York Knicks':'뉴욕닉스',
             'Orlando Magic':'올랜매직',
             'Portland Trail Blazers':'포틀트레',
             'Dallas Mavericks':'댈러매버',
             'Detroit Pistons':'디트피스',
             'Houston Rockets':'휴스로케'}

# NBA 공홈 선수별 데이터 크롤링
def nba_Player_DataLoad(year):

    fileName = nba_mkdir_path + 'NBA_'+ year +' 시즌별_선수_통합_데이터.xlsx'

    print(fileName)

    wb = openpyxl.Workbook()
    ws = wb.active
    wb.save(fileName)  

    url = f"https://stats.nba.com/stats/leaguedashplayerstats?College=&Conference=&Country=&DateFrom=&DateTo=&Division=&DraftPick=&DraftYear=&GameScope=&GameSegment=&Height=&LastNGames=0&LeagueID=00&Location=&MeasureType=Base&Month=0&OpponentTeamID=0&Outcome=&PORound=0&PaceAdjust=N&PerMode=Totals&Period=0&PlayerExperience=&PlayerPosition=&PlusMinus=N&Rank=N&Season={year}&SeasonSegment=&SeasonType=Regular%20Season&ShotClockRange=&StarterBench=&TeamID=0&VsConference=&VsDivision=&Weight="

    headers = {
        "Referer": f"https://www.nba.com/",            
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36",
    }
    
    row = 1
    with HTMLSession() as s:
        result = s.get(url, headers=headers).text
        json_result = json.loads(result)

        output = json_result['resultSets'][0]

        typeList = output['headers']
        playerList = output['rowSet']

        index = 1
        for type in typeList:

            ## 엑셀 저장
            ws.cell(row, index).value = type
            index += 1

        wb.save(fileName)
        row += 1

        for playerData in playerList:
            
            index = 1
            for player in playerData:

                ## 엑셀 저장
                ws.cell(row, index).value = player
                index += 1

            print(row)
            row += 1
            wb.save(fileName)

# NBA 공홈 팀별 데이터 크롤링
def nba_Team_DataLoad(year):

    fileName = nba_mkdir_path + 'NBA_'+year+' 시즌_팀별_통합_데이터.xlsx'

    print(fileName)

    wb = openpyxl.Workbook()
    ws = wb.active
    wb.save(fileName)  

    url = f"https://stats.nba.com/stats/leaguedashteamstats?Conference=&DateFrom=&DateTo=&Division=&GameScope=&GameSegment=&Height=&LastNGames=0&LeagueID=00&Location=&MeasureType=Base&Month=0&OpponentTeamID=0&Outcome=&PORound=0&PaceAdjust=N&PerMode=Totals&Period=0&PlayerExperience=&PlayerPosition=&PlusMinus=N&Rank=N&Season={year}&SeasonSegment=&SeasonType=Regular%20Season&ShotClockRange=&StarterBench=&TeamID=0&TwoWay=0&VsConference=&VsDivision="

    headers = {
        "Referer": f"https://www.nba.com/",            
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36",
    }
    
    row = 1
    with HTMLSession() as s:
        result = s.get(url, headers=headers).text
        json_result = json.loads(result)

        output = json_result['resultSets'][0]

        typeList = output['headers']
        playerList = output['rowSet']

        index = 1
        for type in typeList:

            ## 엑셀 저장
            ws.cell(row, index).value = type
            index += 1

        wb.save(fileName)
        row += 1

        for playerData in playerList:
            
            index = 1
            for player in playerData:

                ## 엑셀 저장
                ws.cell(row, index).value = player
                index += 1

            print(row)
            row += 1
            wb.save(fileName)

# NBA 선수 스탯 통계
def nba_Player_Analysis():

    yearList = ['2019-20','2020-21','2021-22','2022-23']

    save_fileName = nba_mkdir_path + 'NBA_선수_스탯_분석_데이터.xlsx'
    print(save_fileName)

    save_wb = openpyxl.Workbook()
    save_ws = save_wb.active
    save_ws.append(['선수ID','선수이름', '팀명','2019-20시즌','2020-21시즌','2021-22시즌','2022-23시즌', '총합'])
    save_wb.save(save_fileName) 

    for year in yearList:

        loadfileName = nba_mkdir_path + 'NBA_' + year + ' 시즌별_선수_통합_데이터.xlsx'

        load_wb = load_workbook(filename = loadfileName, data_only=True)
        ws = load_wb[load_wb.sheetnames[0]]

        print(loadfileName)

        book = openpyxl.load_workbook(save_fileName)
        sheet = book.worksheets[0]

        index = 1
        for row in ws.rows:

            # 0=PLAYER_ID, 1=PLAYER_NAME, 2=NICKNAME, 3=TEAM_ID, 4=TEAM_ABBREVIATION, 5=AGE
            # 6=GP, 7=W, 8=L, 9=W_PCT, 10=MIN, 11=FGM, 12=FGA, 13=FG_PCT, 14=FG3M, 15=FG3A, 16=FG3_PCT
            # 17=FTM, 18=FTA, 19=FT_PCT, 20=OREB, 21=DREB, 22=REB, 23=AST, 24=TOV, 25=STL, 26=BLK
            # 27=BLKA, 28=PF, 29=PFD, 30=PTS, 31=PLUS_MINUS, 32=NBA_FANTASY_PTS, 33=DD2, 34=TD3, 35=WNBA_FANTASY_PTS

            if index == 1:
                index += 1
                continue

            _playerID = row[0].value # 플레이어 ID
            _playerName = row[1].value # 플레이어 명
            _teamName = row[4].value # 팀명
            _win = row[7].value # 승수
            _pts = row[30].value # 득점
            _reb = row[22].value # 총 리바운드
            _ast = row[23].value # 어시스트
            _stl = row[25].value # 스틸
            _blk = row[26].value # 블록
            _fga = row[12].value # 야투 시도 횟수
            _fgm = row[11].value # 야투 성공 횟수
            _fta = row[18].value # 자유투 시도 횟수
            _ftm = row[17].value # 자유투 성공 횟수
            _tov = row[24].value # 실책(턴오버)
            _pf = row[28].value # 개인 파울

            # NBA 공식 선수 효율성
            _nbaEfficiency = (_pts + _reb + _ast + _stl + _blk) - ((_fga - _fgm) + (_fta - _ftm) + _tov)

            if year == '2019-20': 
                save_ws.cell(index, 1).value = _playerID
                save_ws.cell(index, 2).value = _playerName
                save_ws.cell(index, 3).value = _teamName
                save_ws.cell(index, 4).value = _nbaEfficiency
            elif year == '2020-21':
                saveNumber = 1
                for saveRow in sheet.rows:
                    if saveRow[0].value == _playerID:
                        save_ws.cell(saveNumber, 5).value = _nbaEfficiency
                        break
                    
                    saveNumber += 1
            elif year == '2021-22':
                saveNumber = 1
                for saveRow in sheet.rows:
                    if saveRow[0].value == _playerID:
                        save_ws.cell(saveNumber, 6).value = _nbaEfficiency
                        break

                    saveNumber += 1
            elif year == '2022-23':
                saveNumber = 1
                for saveRow in sheet.rows:
                    if saveRow[0].value == _playerID:
                        save_ws.cell(saveNumber, 7).value = _nbaEfficiency
                        break

                    saveNumber += 1

            index += 1
            save_wb.save(save_fileName)  

    save_fileName = nba_mkdir_path + 'NBA_선수_스탯_분석_데이터.xlsx'

    #엑셀파일 불러오기
    wb = load_workbook(save_fileName)
    ws = wb.active

    delList = []

    index = 1
    for row in ws.rows:
        if index == 1:
            index += 1
            continue

        if row[3].value == None or row[4].value == None or row[5].value == None or row[6].value == None:
            # ws.delete_rows(index)
            index += 1
            continue
            # delList.append(index)

        # 19~20 - 0.1 / 20~21 - 0.2 / 21~22 - 0.3 / 22~23 - 0.4
        result = (row[3].value*0.1) + (row[4].value*0.2) + (row[5].value*0.3) +(row[6].value*0.4)

        ws.cell(index, 8).value = result

        index += 1
    
    wb.save(save_fileName)

    data=pd.read_excel(save_fileName) #원본 엑셀 파일 
    data=data.sort_values(by='총합', ascending=False) # 리뷰개수로 내림차순 정렬

    with pd.ExcelWriter(save_fileName) as writer:
        data.to_excel(writer, sheet_name="sheet0",index=False) # 그대로 저장

# NBA 팀 스탯 통계
def nba_Team_Analysis():

    yearList = ['2019-20','2020-21','2021-22','2022-23']

    save_fileName = nba_mkdir_path + 'NBA_팀_스탯_분석_데이터.xlsx'
    print(save_fileName)

    save_wb = openpyxl.Workbook()
    save_ws = save_wb.active
    save_ws.append(['팀ID','팀명','2019-20시즌','2020-21시즌','2021-22시즌','2022-23시즌', '총합'])
    save_wb.save(save_fileName) 

    for year in yearList:

        loadfileName = nba_mkdir_path + 'NBA_' + year + ' 시즌_팀별_통합_데이터.xlsx'

        load_wb = load_workbook(filename = loadfileName, data_only=True)
        ws = load_wb[load_wb.sheetnames[0]]

        print(loadfileName)

        book = openpyxl.load_workbook(save_fileName)
        sheet = book.worksheets[0]

        index = 1
        for row in ws.rows:

            # 0=TEAM_ID, 1=TEAM_NAME 2=GP, 3=W, 4=L, 5=W_PCT, 6=MIN, 7=FGM, 8=FGA, 9=FG_PCT
            # 10=FG3M, 11=FG3A, 12=FG3_PCT, 13=FTM, 14=FTA, 15= FT_PCT, 16=OREB, 17=DREB, 18=REB, 19=AST, 20=TOV, 21=STL
            # 22=BLK 23=BLKA, 24=PF, 25=PFD, 26=PTS, 27=PLUS_MINUS

            if index == 1:
                index += 1
                continue

            _teamID = row[0].value # 팀 ID
            _teamName = row[1].value # 팀명
            _pts = row[23].value # 득점
            _oreb = row[16].value # 공격 리바운드
            _dreb = row[17].value # 수비 리바운드
            _reb = row[15].value # 총 리바운드
            _fgm = row[7].value # 야투 성공 횟수
            _fga = row[8].value # 야투 시도 횟수
            _fta = row[14].value # 자유투 시도 횟수
            _ftm = row[13].value # 자유투 성공 횟수
            _stl = row[21].value # 스틸
            _ast = row[19].value # 어시스트
            _blk = row[22].value # 블록
            _tov = row[20].value # 실책(턴오버)
            _pf = row[24].value # 개인 파울

            # NBA 팀 스코어 (2000은 음수를 없애기 위함)
            _nbaTeamScore = (_pts + 0.4*_fgm - 0.7*_fga - 0.4*(_fta - _ftm) + 0.7*_oreb + 0.3*_dreb + _stl +0.7*_ast + 0.7*_blk - 0.4*_pf - _tov) + 2000

            if year == '2019-20': 
                save_ws.cell(index, 1).value = _teamID
                save_ws.cell(index, 2).value = _teamName
                save_ws.cell(index, 3).value = _nbaTeamScore
            elif year == '2020-21':
                saveNumber = 1
                for saveRow in sheet.rows:
                    if saveRow[0].value == _teamID:
                        save_ws.cell(saveNumber, 4).value = _nbaTeamScore
                        break
                    
                    saveNumber += 1
            elif year == '2021-22':
                saveNumber = 1
                for saveRow in sheet.rows:
                    if saveRow[0].value == _teamID:
                        save_ws.cell(saveNumber, 5).value = _nbaTeamScore
                        break

                    saveNumber += 1
            elif year == '2022-23':
                saveNumber = 1
                for saveRow in sheet.rows:
                    if saveRow[0].value == _teamID:
                        save_ws.cell(saveNumber, 6).value = _nbaTeamScore
                        break

                    saveNumber += 1

            index += 1
            save_wb.save(save_fileName)  

    save_fileName = nba_mkdir_path + 'NBA_팀_스탯_분석_데이터.xlsx'

    #엑셀파일 불러오기
    wb = load_workbook(save_fileName)
    ws = wb.active

    index = 1
    for row in ws.rows:
        if index == 1:
            index += 1
            continue

        # 19~20 - 0.1 / 20~21 - 0.2 / 21~22 - 0.3 / 22~23 - 0.4
        result = (row[2].value*0.1) + (row[3].value*0.2) + (row[4].value*0.3) +(row[5].value*0.4)

        ws.cell(index, 7).value = result

        index += 1
    
    wb.save(save_fileName)

    data=pd.read_excel(save_fileName) #원본 엑셀 파일 
    data=data.sort_values(by='총합', ascending=False) # 리뷰개수로 내림차순 정렬

    with pd.ExcelWriter(save_fileName) as writer:
        data.to_excel(writer, sheet_name="sheet0",index=False) # 그대로 저장

# NBA 2019~2023 상대 전적 (홈, 원정 동일 위치) - typeText 'Home', 'Away'
def nba_Betman_vs_Score(typeText, homeName, awayName):

    yearList = [2019,2020,2021,2022,2023]

    teamScoreList = []

    for year in yearList:

        loadfileName = nba_mkdir_path + '배트맨_NBA_' + str(year) + '년도_데이터.xlsx'

        wb = load_workbook(filename = loadfileName, data_only=True)
        ws = wb[wb.sheetnames[0]]

        for row in ws.rows:

            # 0='No', 1='종류', 2='날짜', 3='리그 이름', 4='홈 네임', 5='원정 네임', 6='승리 배당'
            # 7='무 배당', 8='패배 배당', 9='홈 핸디', 10='무 핸디', 11='원정 핸디', 12='경기 결과', 13='경기 결과 점수'

            _date = str(row[2].value)

            _homeName = str(row[4].value)
            _awayName = str(row[5].value)

            _matchResult = str(row[13].value)

            if _homeName != homeName or _awayName != awayName: continue

            if typeText == 'Home':
                teamScoreList.append(_matchResult.split(':')[0])
            elif typeText == 'Away':
                teamScoreList.append(_matchResult.split(':')[1])
                    
    teamScore = 0
    for score in teamScoreList:
        teamScore += int(score)

    return teamScore/len(teamScoreList)

# NBA 2022~2023 최근 전적 (홈, 원정 동일 위치) - typeText 'Home', 'Away'
def nba_Betman_RecentScore(typeText, teamName):

    yearList = [2022,2023]

    teamScoreList = []

    for year in yearList:

        loadfileName = nba_mkdir_path + '배트맨_NBA_' + str(year) + '년도_데이터.xlsx'

        wb = load_workbook(filename = loadfileName, data_only=True)
        ws = wb[wb.sheetnames[0]]

        for row in ws.rows:

            # 0='No', 1='종류', 2='날짜', 3='리그 이름', 4='홈 네임', 5='원정 네임', 6='승리 배당'
            # 7='무 배당', 8='패배 배당', 9='홈 핸디', 10='무 핸디', 11='원정 핸디', 12='경기 결과', 13='경기 결과 점수'

            _date = str(row[2].value)

            _homeName = str(row[4].value)
            _awayName = str(row[5].value)
            _matchResult = str(row[13].value)

            if typeText == 'Home':
                if _homeName == teamName: 
                    teamScoreList.append(_matchResult.split(':')[0])
            elif typeText == 'Away':
                if _awayName == teamName: 
                    teamScoreList.append(_matchResult.split(':')[1])
                    
    teamScore = 0
    for score in teamScoreList:
        teamScore += int(score)

    return teamScore/len(teamScoreList)

# 배트맨 NBA 데이터 크롤링
def nba_Betman_DataLoad(year):

    fileName = nba_mkdir_path + '배트맨_NBA_' + str(year) + '년도_데이터.xlsx'

    print(fileName)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['No', '종류', '날짜', '리그 이름', '홈 네임', '원정 네임','승리 배당','무 배당','패배 배당','홈 핸디','무 핸디','원정 핸디','경기 결과','경기 결과 점수'])

    wb.save(fileName)  

    year = str(year)[2:]

    url = "https://www.betman.co.kr/buyPsblGame/gameInfoInq.do"

    # 210001 = 21년도 / 0001 -> 1주차 경기
    dataNumbering = int(year+'0001')

    row = 2
    for i in range(200):

        print('데이터 주차 : ' + str(dataNumbering))

        headers = {
            "Referer": f"https://www.betman.co.kr/main/mainPage/gamebuy/gameSlipIFR.do?gmId=G101&gmTs={str(dataNumbering)}&gameDivCd=C&isIFR=Y",            
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36",
            'accept': '*/*',
            'accept-encoding': 'gzip, deflate, br',
            'accept-language': 'ko',
            'Content-Type': 'application/json'
        }

        data = {
            "gmId" : "G101",
            "gmTs" : str(dataNumbering),
            "gameYear" : "",
            "_sbmInfo" : {"_sbmInfo" : {"debugMode": "false"}}
        }

        with HTMLSession() as s:
            result = s.post(url, headers=headers, data=json.dumps(data)).text
            json_result = json.loads(result)

            try:
                output = json_result['compSchedules']['datas']
            except:
                print(str(dataNumbering) + ' : 데이터가 없습니다.')
                break

            dataNumbering += 1

            for li in output:

                _type = li[0] # 스포츠 종류 (SC=축구/BK=농구/VL=배구)

                # 경기 날짜, 시간
                date = str(li[3])[:len(str(li[3]))-3] # 13자리에서 뒤에 3자리 빼고 변환(파이썬에서는 그렇게 해야 변환되서)
                _matchDate = datetime.datetime.fromtimestamp(int(date)).strftime('%Y-%m-%d %H:%M:%S')

                _leagueName = li[7] # 리그 이름
                _homeName = li[14] # 홈 네임
                _awayName = li[15] # 원정 네임
                _winRate = li[16] # 승리 배당률
                _drawRate = li[17] # 무승부 배당률
                _lossRate = li[18] # 패배 배당률

                _homeHandy = li[20] # 홈 핸디
                _drawHandy = li[21] # 무 핸디
                _awayHandy = li[22] # 원정 핸디

                _result = li[28] # 경기 결과 - 홈 0 / 무 1 / 원정 2

                _matchResult = li[33] # 경기 결과 점수

                if _type != 'BK' or _matchResult == None or _homeHandy != 0.0: 
                    continue

                ## 엑셀 저장
                ws.cell(row, 1).value = row-1
                ws.cell(row, 2).value = _type
                ws.cell(row, 3).value = _matchDate
                ws.cell(row, 4).value = _leagueName
                ws.cell(row, 5).value = _homeName
                ws.cell(row, 6).value = _awayName
                ws.cell(row, 7).value = _winRate
                ws.cell(row, 8).value = _drawRate
                ws.cell(row, 9).value = _lossRate
                ws.cell(row, 10).value = _homeHandy
                ws.cell(row, 11).value = _drawHandy
                ws.cell(row, 12).value = _awayHandy
                ws.cell(row, 13).value = _result
                ws.cell(row, 14).value = _matchResult
                row += 1

    wb.save(fileName)

# NBA 공홈 금일 팀, 선수 라인업 크롤링해서 분석
def nba_DailyLineup_Analysis(daily):

    playerFileName = nba_mkdir_path + 'NBA_선수_스탯_분석_데이터.xlsx'
    teamFileName = nba_mkdir_path + 'NBA_팀_스탯_분석_데이터.xlsx'

    player_wb = load_workbook(filename = playerFileName, data_only=True)
    player_ws = player_wb[player_wb.sheetnames[0]]

    team_wb = load_workbook(filename = teamFileName, data_only=True)
    team_ws = team_wb[team_wb.sheetnames[0]]

    url = f"https://stats.nba.com/js/data/leaders/00_daily_lineups_{daily}.json"

    headers = {
        "Referer": "https://www.nba.com/",            
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36",
    }
    
    with HTMLSession() as s:
        result = s.get(url, headers=headers).text
        json_result = json.loads(result)

        output = json_result['games']

        for games in output:

            homePlayerIds = []
            awayPlayerIds = []

            homeTeamId = games['homeTeam']['teamId'] 
            awayTeamId = games['awayTeam']['teamId'] 

            for home in games['homeTeam']['players']:
                homePlayerIds.append(home['personId'])

            for away in games['awayTeam']['players']:
                awayPlayerIds.append(away['personId'])

            # 플레이어
            homeIndex = 1
            awayIndex = 1

            homeResult = 0
            awayResult = 0
            for row in player_ws.rows:
                for id in homePlayerIds:
                    if row[0].value == id and row[7].value != None:
                        homeResult += row[7].value
                        homeIndex += 1
                        break

                for id in awayPlayerIds:
                    if row[0].value == id and row[7].value != None:
                        awayResult += row[7].value
                        awayIndex += 1
                        break
            
            # 팀
            homeTeamResult = 1
            awayTeamResult = 1

            homeTeamName = ''
            awayTeamName = ''
            for row in team_ws.rows:
                if row[0].value == homeTeamId:
                    homeTeamName = row[1].value
                    homeTeamResult = row[6].value

                if row[0].value == awayTeamId:
                    awayTeamName = row[1].value
                    awayTeamResult = row[6].value

                
            # (상대전적x0.65) + (최근전적x0.35)
            vsScore = nba_Betman_vs_Score('Home',TEAM_EN_TO[homeTeamName],TEAM_EN_TO[awayTeamName])
            recentScore = nba_Betman_RecentScore('Home',TEAM_EN_TO[homeTeamName])

            homeScore = (vsScore*0.65) + (recentScore*0.35)

            # (상대전적x0.65) + (최근전적x0.35)
            vsScore = nba_Betman_vs_Score('Away',TEAM_EN_TO[homeTeamName],TEAM_EN_TO[awayTeamName])
            recentScore = nba_Betman_RecentScore('Away',TEAM_EN_TO[awayTeamName])

            awayScore = (vsScore*0.65) + (recentScore*0.35)

            homeResult = (homeResult / homeIndex) + homeTeamResult
            awayResult = (awayResult / awayIndex) + awayTeamResult

            if homeResult > awayResult:
                changeHomeScore = int(homeScore)
                changeAwayScore = changeHomeScore - 10
            else:
                changeAwayScore = int(awayScore)
                changeHomeScore = changeAwayScore - 10
                
            print(TEAM_EN_TO[homeTeamName] + ' : ' + TEAM_EN_TO[awayTeamName])
            print('기본 -> ' + str(int(homeScore)) + ' : ' + str(int(awayScore)))
            print('스탯 점수 추가 후 -> ' + str(changeHomeScore) + ' : ' + str(changeAwayScore))
            print('■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■')

if __name__ == '__main__':

    # ■■■■■■■■■■ 매일 데이터 업데이트 ■■■■■■■■■■

    # nba_Player_DataLoad('2022-23')
    # nba_Team_DataLoad('2022-23')
    # nba_Betman_DataLoad(2023)

    # nba_Player_Analysis()
    # nba_Team_Analysis()

    # ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■

    # 미국 날짜로 금일 -1일
    nba_DailyLineup_Analysis('20230216')










   
