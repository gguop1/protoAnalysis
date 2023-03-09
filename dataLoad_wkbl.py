from requests_html import HTMLSession
import requests
import json
import urllib
import openpyxl
import pandas as pd
import time
from wordpress_xmlrpc import Client
from wordpress_xmlrpc import WordPressPost
from wordpress_xmlrpc.methods import posts
from openpyxl import load_workbook
from datetime import datetime
import os.path   
import random
import pprint
import datetime
from collections import Counter
import numpy as np
import pprint
from bs4 import BeautifulSoup

wkbl_mkdir_path = 'D:/pythonAppCode/ProtoAnalysis/WKBL/'

TEAM_ID = {'KB스타즈':'01',
           '삼성생명':'03',
           '우리은행':'05',
           '신한은행':'07',
           '하나원큐':'09',
           'BNK썸':'11'}

# WKBL 공홈 팀별 데이터 크롤링 2020 ~ 2023
def wkbl_Team_DataLoad(year):

    fileName = wkbl_mkdir_path + 'WKBL_팀_'+ str(year) +'시즌별_통합_데이터.xlsx'

    print(fileName)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['순위','팀명','경기수','득점','실점','리바운드', '어시스트','스틸','블록','3점슛','2점슛','자유투','3점슛성공률','2점슛성공률','자유투성공률'])
    wb.save(fileName)  

    if year == 2023: year = '043'
    elif year == 2022: year = '042'
    elif year == 2021: year = '041'
    elif year == 2020: year = '040'

    url = f"https://www.wkbl.or.kr/game/ajax/ajax_part_team_rank.asp"

    headers = {
        'referer':'https://www.wkbl.or.kr/game/part_team_rank.asp',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36',
    }

    data = {
        'season_gu': year,
        'opart': '1',
        'opart_pre': '1',
        'sort': 'desc',
    }

    with HTMLSession() as s:
        req = s.post(url, headers=headers, data=data)
        soup = BeautifulSoup(req.content, 'html.parser')

        row = 2
        sell = 1
        for td in soup.find_all('td'):

            if sell == 16:
                sell = 1
                row += 1

            ws.cell(row, sell).value = td.text.replace(' ','')
            sell += 1
        
        wb.save(fileName)
 
# WKBL 공홈 선수별 데이터 크롤링 2020 ~ 2023
def wkbl_Player_DataLoad(year):

    fileName = wkbl_mkdir_path + 'WKBL_선수_'+ str(year) +'시즌별_통합_데이터.xlsx'

    print(fileName)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['팀ID','선수','경기수','출전시간','2점슛 성공-시도','3점슛 성공-시도','자유투 성공-시도','공격리바운드','수비리바운드','리바운드','어시스트','스틸','블록슛','턴오버','파울','총 득점'])
    wb.save(fileName)  

    if year == 2023: year = '043'
    elif year == 2022: year = '042'
    elif year == 2021: year = '041'
    elif year == 2020: year = '040'

    teamIDList = ['01','03','05','07','09','11']

    row = 2
    sell = 1
    for teamID in teamIDList:

        url = "https://www.wkbl.or.kr/game/ajax/ajax_team_total_record.asp"

        headers = {
            'referer':'https://www.wkbl.or.kr/game/team_total_record.asp',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36',
        }

        data = {
            'season_gu': year,
            'team_code': teamID,
        }

        check = False
        checkIndex = 1
        with HTMLSession() as s:
            req = s.post(url, headers=headers, data=data)
            soup = BeautifulSoup(req.content, 'html.parser')

            for td in soup.find_all('td'):

                if td.text == '팀평균':
                    check = True
                    continue

                if check == False: continue

                if check == True and checkIndex < 15:
                    checkIndex += 1
                    continue

                if td.text == '팀합계': break

                if sell == 17:
                    sell = 1
                    row += 1

                if sell == 1:
                    ws.cell(row, 1).value = teamID
                    sell += 1

                ws.cell(row, sell).value = td.text
                sell += 1
            
            wb.save(fileName)

# WKBL 팀 스탯 통계
def wkbl_Team_Analysis():

    yearList = ['2020','2021','2022','2023']

    save_fileName = wkbl_mkdir_path + 'WKBL_팀_스탯_분석_데이터.xlsx'
    print(save_fileName)

    save_wb = openpyxl.Workbook()
    save_ws = save_wb.active
    save_ws.append(['팀ID','팀 명','2019-20시즌','2020-21시즌','2021-22시즌','2022-23시즌', '총합'])
    save_wb.save(save_fileName) 

    for year in yearList:

        loadfileName = wkbl_mkdir_path + 'WKBL_팀_' + year + '시즌별_통합_데이터.xlsx'

        load_wb = load_workbook(filename = loadfileName, data_only=True)
        ws = load_wb[load_wb.sheetnames[0]]

        print(loadfileName)

        book = openpyxl.load_workbook(save_fileName)
        sheet = book.worksheets[0]

        index = 1
        for row in ws.rows:

            if index == 1:
                index += 1
                continue

            _teamName = row[1].value.replace(' ','') # 팀명

            if _teamName == '하나은행': _teamName = '하나원큐'

            _pts = float(row[3].value) # 득점
            _reb = float(row[5].value) # 총 리바운드
            _fgm = float(row[9].value) + float(row[10].value) # 야투 성공 횟수
            _ftm = float(row[11].value) # 자유투 성공 횟수
            _stl = float(row[7].value) # 스틸
            _ast = float(row[6].value) # 어시스트
            _blk = float(row[8].value) # 블록
            _tov = float(row[4].value) # 실책(턴오버)

            # KBL 팀 스코어
            _kblTeamScore = _pts + 0.4*_fgm - 0.4*_ftm + _reb + _stl +0.7*_ast + 0.7*_blk - _tov

            if year == '2020': 
                save_ws.cell(index, 1).value = TEAM_ID[_teamName]
                save_ws.cell(index, 2).value = _teamName
                save_ws.cell(index, 3).value = _kblTeamScore
            elif year == '2021':
                saveNumber = 1
                checkSave = False
                for saveRow in sheet.rows:
                    if saveRow[1].value == _teamName:
                        save_ws.cell(saveNumber, 4).value = _kblTeamScore
                        checkSave = True
                        break
                    saveNumber += 1

                if checkSave == False:
                    save_ws.append([TEAM_ID[_teamName],_teamName,0,_kblTeamScore])
            elif year == '2022':
                saveNumber = 1
                checkSave = False
                for saveRow in sheet.rows:
                    if saveRow[1].value == _teamName:
                        save_ws.cell(saveNumber, 5).value = _kblTeamScore
                        checkSave = True
                        break
                    saveNumber += 1

                if checkSave == False:
                    save_ws.append([TEAM_ID[_teamName],_teamName,0,0,_kblTeamScore])
            elif year == '2023':
                saveNumber = 1
                checkSave = False
                for saveRow in sheet.rows:
                    if saveRow[1].value == _teamName:
                        save_ws.cell(saveNumber, 6).value = _kblTeamScore
                        checkSave = True
                        break
                    saveNumber += 1
                if checkSave == False:
                    save_ws.append([TEAM_ID[_teamName],_teamName,0,0,0,_kblTeamScore])

            index += 1
            save_wb.save(save_fileName)  

    #엑셀파일 불러오기
    wb = load_workbook(save_fileName)
    ws = wb.active

    index = 1
    for row in ws.rows:
        if index == 1:
            index += 1
            continue

        dataInCheck = 4

        season20 = row[2].value
        if season20 == None: 
            season20 = 0
            dataInCheck -= 1
        season21 = row[3].value
        if season21 == None: 
            season21 = 0
            dataInCheck -= 1
        season22 = row[4].value
        if season22 == None: 
            season22 = 0
            dataInCheck -= 1
        season23 = row[5].value
        if season23 == None: 
            season23 = 0
            dataInCheck -= 1

        # 19~20 - 0.1 / 20~21 - 0.2 / 21~22 - 0.3 / 22~23 - 0.4
        result = ((season20*0.1) + (season21*0.2) + (season22*0.3) + (season23*4)) / dataInCheck

        ws.cell(index, 7).value = result

        index += 1
    
    wb.save(save_fileName)

# WKBL 선수 스탯 통계
def wkbl_Player_Analysis():

    yearList = ['2020','2021','2022','2023']

    save_fileName = wkbl_mkdir_path + 'WKBL_선수_스탯_분석_데이터.xlsx'
    print(save_fileName)

    save_wb = openpyxl.Workbook()
    save_ws = save_wb.active
    save_ws.append(['팀ID','선수 명','2019-20시즌','2020-21시즌','2021-22시즌','2022-23시즌', '총합'])
    save_wb.save(save_fileName) 

    for year in yearList:

        loadfileName = wkbl_mkdir_path + 'WKBL_선수_' + year + '시즌별_통합_데이터.xlsx'

        load_wb = load_workbook(filename = loadfileName, data_only=True)
        ws = load_wb[load_wb.sheetnames[0]]

        print(loadfileName)

        book = openpyxl.load_workbook(save_fileName)
        sheet = book.worksheets[0]

        index = 1
        for row in ws.rows:

            if index == 1:
                index += 1
                continue

            _teamID = row[0].value # 팀 ID
            _PlayerName = row[1].value # 플레이어 명

            _pts = float(row[15].value) # 득점
            _oreb = float(row[7].value) # 공격 리바운드
            _dreb = float(row[8].value) # 수비 리바운드

            _pt1_fail = int(row[6].value.split('-')[1]) - int(row[6].value.split('-')[0]) # 자유투 실패 횟수
            _pt2_fail = int(row[4].value.split('-')[1]) - int(row[4].value.split('-')[0]) # 2점슛 실패 횟수
            _pt3_fail = int(row[5].value.split('-')[1]) - int(row[5].value.split('-')[0]) # 3점슛 실패 횟수

            _stl = float(row[11].value) # 스틸
            _ast = float(row[10].value) # 어시스트
            _blk = float(row[12].value) # 블록
            _tov = float(row[13].value) # 실책(턴오버)
            _playMin = float(row[3].value.strip().split(':')[0]) # 출전시간(분)

            # 가산법
            _kblEfficiencyAdd = (_pts+_stl+_blk+_dreb)*1.0 + (_oreb+_ast)*1.5 + _playMin / 4
            _kblEfficiencyMinus = _tov*1.5 + _pt2_fail*1.0 +_pt3_fail*0.9 + _pt1_fail*0.8

            # KBL 플레이어 가성비
            _kblPlayerEfficiency = _kblEfficiencyAdd - _kblEfficiencyMinus

            if year == '2020': 
                save_ws.cell(index, 1).value = _teamID
                save_ws.cell(index, 2).value = _PlayerName
                save_ws.cell(index, 3).value = _kblPlayerEfficiency
            elif year == '2021':
                saveNumber = 1
                checkSave = False
                for saveRow in sheet.rows:
                    if saveRow[0].value == _teamID and saveRow[1].value == _PlayerName:
                        save_ws.cell(saveNumber, 4).value = _kblPlayerEfficiency
                        checkSave = True
                        break
                    saveNumber += 1

                if checkSave == False:
                    save_ws.append([_teamID,_PlayerName,0,_kblPlayerEfficiency])
            elif year == '2022':
                saveNumber = 1
                checkSave = False
                for saveRow in sheet.rows:
                    if saveRow[0].value == _teamID and saveRow[1].value == _PlayerName:
                        save_ws.cell(saveNumber, 5).value = _kblPlayerEfficiency
                        checkSave = True
                        break
                    saveNumber += 1

                if checkSave == False:
                    save_ws.append([_teamID,_PlayerName,0,0,_kblPlayerEfficiency])
            elif year == '2023':
                saveNumber = 1
                checkSave = False
                for saveRow in sheet.rows:
                    if saveRow[0].value == _teamID and saveRow[1].value == _PlayerName:
                        save_ws.cell(saveNumber, 6).value = _kblPlayerEfficiency
                        checkSave = True
                        break
                    saveNumber += 1
                if checkSave == False:
                    save_ws.append([_teamID,_PlayerName,0,0,0,_kblPlayerEfficiency])


            index += 1
            save_wb.save(save_fileName)  

    #엑셀파일 불러오기
    wb = load_workbook(save_fileName)
    ws = wb.active

    index = 1
    for row in ws.rows:
        if index == 1:
            index += 1
            continue

        dataInCheck = 4

        season20 = row[2].value
        if season20 == None: 
            season20 = 0
            dataInCheck -= 1
        season21 = row[3].value
        if season21 == None: 
            season21 = 0
            dataInCheck -= 1
        season22 = row[4].value
        if season22 == None: 
            season22 = 0
            dataInCheck -= 1
        season23 = row[5].value
        if season23 == None: 
            season23 = 0
            dataInCheck -= 1

        # 19~20 - 0.1 / 20~21 - 0.2 / 21~22 - 0.3 / 22~23 - 0.4
        result = ((season20*0.1) + (season21*0.2) + (season22*0.3) + (season23*0.4)) / dataInCheck

        ws.cell(index, 7).value = result

        index += 1
    
    wb.save(save_fileName)

# WKBL 전반전 점수 통계
def wkbl_FirstHalf_Data(homeName, awayName):

    homeScore = 0
    awayScore = 0

    url = 'https://www.wkbl.or.kr/game/ajax/ajax_report.asp'

    headers = {
        'referer': 'https://www.wkbl.or.kr/game/report.asp',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36'
    }

    data = {
        'tcode1': TEAM_ID[homeName],
        'tcode2': TEAM_ID[awayName]
    }
    
    with HTMLSession() as s:
        req = s.post(url, headers=headers, data=data)
        soup = BeautifulSoup(req.content, 'html.parser')

        # 1 - 날짜 / 2 - 경기번호 / 3 - 홈팀 / 4 - 원정팀 / 5 - 경기장
        # 6 - 팀 이름★ / 7 - 1쿼터★ / 8 - 2쿼터★ / 9 - 3쿼터 / 10 - 4쿼터
        # 11 - 연장 / 12 - 총 스코어 / 13 - 승리팀 / 14 - 팀 이름★ / 15 - 1쿼터★ 
        # 16 - 2쿼터★ / 17 - 3쿼터 / 18 - 4쿼터 / 19 - 연장

        index = 1
        row = 0
        teamCheck = 'home'
        for td in soup.find_all('td'):

            if index == 20:
                index = 1

            if index == 6 and td.text == homeName: teamCheck = 'home'
            elif index == 6 and td.text == awayName: teamCheck = 'away'

            if index == 7 or index == 8:
                if teamCheck == 'home':
                    homeScore += int(td.text)
                    row += 1
                elif teamCheck == 'away':
                    awayScore += int(td.text)

            if index == 14 and td.text == homeName: teamCheck = 'home'
            elif index == 14 and td.text == awayName: teamCheck = 'away'

            if index == 15 or index == 16:
                if teamCheck == 'home':
                    homeScore += int(td.text)
                    row += 1
                elif teamCheck == 'away':
                    awayScore += int(td.text)

            index += 1

    scoreList = []
    scoreList.append(int(homeScore/(row/2)))
    scoreList.append(int(awayScore/(row/2)))

    print(scoreList)

    return scoreList

# 배트맨 WKBL 2019~2023 상대 전적 (홈, 원정 동일 위치) - typeText 'Home', 'Away'
def betman_vs_Score(typeText, homeID, awayID):

    yearList = [2019,2020,2021,2022,2023]

    teamScoreList = []

    for year in yearList:

        loadfileName = wkbl_mkdir_path + '배트맨_WKBL_' + str(year) + '년도_데이터.xlsx'

        wb = load_workbook(filename = loadfileName, data_only=True)
        ws = wb[wb.sheetnames[0]]

        for row in ws.rows:

            _date = str(row[2].value)

            _homeID = str(row[4].value)
            _awayID = str(row[5].value)

            _matchResult = str(row[15].value)

            if _homeID != homeID or _awayID != awayID: continue

            if typeText == 'Home':
                teamScoreList.append(_matchResult.split(':')[0])
            elif typeText == 'Away':
                teamScoreList.append(_matchResult.split(':')[1])
                    
    teamScore = 0
    for score in teamScoreList:
        teamScore += int(score)

    return teamScore/len(teamScoreList)

# 배트맨 WKBL 2022~2023 최근 전적 (홈, 원정 동일 위치) - typeText 'Home', 'Away'
def betman_RecentScore(typeText, teamID):

    yearList = [2022,2023]

    teamScoreList = []

    for year in yearList:

        loadfileName = wkbl_mkdir_path + '배트맨_WKBL_' + str(year) + '년도_데이터.xlsx'

        wb = load_workbook(filename = loadfileName, data_only=True)
        ws = wb[wb.sheetnames[0]]

        for row in ws.rows:

            # 0='No', 1='종류', 2='날짜', 3='리그 이름', 4='홈 네임', 5='원정 네임', 6='승리 배당'
            # 7='무 배당', 8='패배 배당', 9='홈 핸디', 10='무 핸디', 11='원정 핸디', 12='경기 결과', 13='경기 결과 점수'

            _date = str(row[2].value)

            _homeID = str(row[4].value)
            _awayID = str(row[5].value)
            _matchResult = str(row[15].value)

            if typeText == 'Home':
                if _homeID == teamID: 
                    teamScoreList.append(_matchResult.split(':')[0])
            elif typeText == 'Away':
                if _awayID == teamID: 
                    teamScoreList.append(_matchResult.split(':')[1])
                    
    teamScore = 0
    for score in teamScoreList:
        teamScore += int(score)

    return teamScore/len(teamScoreList)

# 배트맨 WKBL 2019~2023 상대 점수차 (홈, 원정 동일 위치)
def betman_Difference_Score(homeID, awayID):

    yearList = [2019,2020,2021,2022,2023]

    teamScoreList = []

    for year in yearList:

        loadfileName = wkbl_mkdir_path + '배트맨_WKBL_' + str(year) + '년도_데이터.xlsx'

        wb = load_workbook(filename = loadfileName, data_only=True)
        ws = wb[wb.sheetnames[0]]

        for row in ws.rows:

            # 0='No', 1='종류', 2='날짜', 3='리그 이름', 4='홈 네임', 5='원정 네임', 6='승리 배당'
            # 7='무 배당', 8='패배 배당', 9='홈 핸디', 10='무 핸디', 11='원정 핸디', 12='경기 결과', 13='경기 결과 점수'

            _homeID = str(row[4].value)
            _awayID = str(row[5].value)

            _matchResult = str(row[15].value)

            if _homeID != homeID or _awayID != awayID: continue

            result = int(_matchResult.split(':')[0]) - int(_matchResult.split(':')[1])

            if result < 0: teamScoreList.append(-result)
            else: teamScoreList.append(result)

                    
    teamScore = 0
    for score in teamScoreList:
        teamScore += int(score)

    return teamScore/len(teamScoreList)

# 배트맨 WKBL 데이터 크롤링
def betman_WKBL_DataLoad(year):

    fileName = wkbl_mkdir_path + '배트맨_WKBL_' + str(year) + '년도_데이터.xlsx'

    print(fileName)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['No', '종류', '날짜', '리그 이름','홈팀 ID','원정팀 ID','홈 네임', '원정 네임','승리 배당','무 배당','패배 배당','홈 핸디','무 핸디','원정 핸디','경기 결과','경기 결과 점수'])

    wb.save(fileName)  

    year = str(year)[2:]

    url = "https://www.betman.co.kr/buyPsblGame/gameInfoInq.do"

    # 210001 = 21년도 / 0001 -> 1주차 경기
    dataNumbering = int(year+'0001')

    row = 2
    for i in range(200):

        print('데이터 주차 : ' + str(dataNumbering))

        headers = {
            "Referer": f"https://www.betman.co.kr/main/mainPage/gamebuy/gameSlipIFR.do?gmId=G101&gmTs={str(dataNumbering)}&gameDivCd=C&isIFR=Y",            
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36",
            'accept': '*/*',
            'accept-encoding': 'gzip, deflate, br',
            'accept-language': 'ko',
            'Content-Type': 'application/json'
        }

        data = {
            "gmId" : "G101",
            "gmTs" : str(dataNumbering),
            "gameYear" : "",
            "_sbmInfo" : {"_sbmInfo" : {"debugMode": "false"}}
        }

        with HTMLSession() as s:
            result = s.post(url, headers=headers, data=json.dumps(data)).text
            json_result = json.loads(result)

            try:
                output = json_result['compSchedules']['datas']
            except:
                print(str(dataNumbering) + ' : 데이터가 없습니다.')
                break

            dataNumbering += 1

            for li in output:

                _type = li[0] # 스포츠 종류 (SC=축구/BK=농구/VL=배구)

                # 경기 날짜, 시간
                date = str(li[3])[:len(str(li[3]))-3] # 13자리에서 뒤에 3자리 빼고 변환(파이썬에서는 그렇게 해야 변환되서)
                _matchDate = datetime.datetime.fromtimestamp(int(date)).strftime('%Y-%m-%d %H:%M:%S')

                _leagueName = li[7] # 리그 이름
                _homeID = li[12] # 홈팀 ID
                _awayID = li[13] # 원정팀 ID
                _homeName = li[14] # 홈 네임
                _awayName = li[15] # 원정 네임
                _winRate = li[16] # 승리 배당률
                _drawRate = li[17] # 무승부 배당률
                _lossRate = li[18] # 패배 배당률

                _homeHandy = li[20] # 홈 핸디
                _drawHandy = li[21] # 무 핸디
                _awayHandy = li[22] # 원정 핸디

                _result = li[28] # 경기 결과 - 홈 0 / 무 1 / 원정 2

                _matchResult = li[33] # 경기 결과 점수

                if _type != 'BK' or _matchResult == None or _homeHandy != 0.0 or _leagueName != 'WKBL': 
                    continue

                ## 엑셀 저장
                ws.cell(row, 1).value = row-1
                ws.cell(row, 2).value = _type
                ws.cell(row, 3).value = _matchDate
                ws.cell(row, 4).value = _leagueName
                ws.cell(row, 5).value = _homeID
                ws.cell(row, 6).value = _awayID
                ws.cell(row, 7).value = _homeName
                ws.cell(row, 8).value = _awayName

                ws.cell(row, 9).value = _winRate
                ws.cell(row, 10).value = _drawRate
                ws.cell(row, 11).value = _lossRate
                ws.cell(row, 12).value = _homeHandy
                ws.cell(row, 13).value = _drawHandy
                ws.cell(row, 14).value = _awayHandy
                ws.cell(row, 15).value = _result
                ws.cell(row, 16).value = _matchResult
                row += 1

    wb.save(fileName)

# 베트맨 금일 라인업 통계
def betman_DailyLineup_Analysis(homeName, awayName):
    
    teamFileName = wkbl_mkdir_path + 'WKBL_팀_스탯_분석_데이터.xlsx'
    playerFileName = wkbl_mkdir_path + 'WKBL_선수_스탯_분석_데이터.xlsx'

    player_wb = load_workbook(filename = playerFileName, data_only=True)
    player_ws = player_wb[player_wb.sheetnames[0]]

    team_wb = load_workbook(filename = teamFileName, data_only=True)
    team_ws = team_wb[team_wb.sheetnames[0]]

    # 플레이어
    homePlayerIndex = 1
    awayPlayerIndex = 1

    homePlayerScore = 0
    awayPlayerScore = 0
    for row in player_ws.rows:
        if row[0].value == TEAM_ID[homeName] and row[5].value != None:
            homePlayerScore += row[6].value
            homePlayerIndex += 1

        if row[0].value == TEAM_ID[awayName] and row[5].value != None:
            awayPlayerScore += row[6].value
            awayPlayerIndex += 1

    # 팀
    homeTeamScore = 1
    awayTeamScore = 1
    for row in team_ws.rows:
        if row[1].value == homeName and row[5].value != None:
            homeTeamScore = int(row[6].value)

        if row[1].value == awayName and row[5].value != None:
            awayTeamScore = int(row[6].value)

        
    # (상대전적x0.65) + (최근전적x0.35)
    vsScore = betman_vs_Score('Home',TEAM_ID[homeName],TEAM_ID[awayName])
    recentScore = betman_RecentScore('Home',TEAM_ID[homeName])

    homeScore = (vsScore*0.65) + (recentScore*0.35)

    # (상대전적x0.65) + (최근전적x0.35)
    vsScore = betman_vs_Score('Away',TEAM_ID[homeName],TEAM_ID[awayName])
    recentScore = betman_RecentScore('Away',TEAM_ID[awayName])

    awayScore = (vsScore*0.65) + (recentScore*0.35)

    homePlayerScore = int((homePlayerScore / homePlayerIndex))
    awayPlayerScore = int((awayPlayerScore / awayPlayerIndex))

    # 홈 팀에 50점 더 추가로 줌
    homeResult = homePlayerScore + 50 + homeTeamScore
    awayResult = awayPlayerScore + awayTeamScore

    gapScore = int(betman_Difference_Score(TEAM_ID[homeName],TEAM_ID[awayName]))

    halfList = wkbl_FirstHalf_Data(homeName, awayName)

    halfGapScore = halfList[0] - halfList[1]

    if halfGapScore < 0: halfGapScore = -halfGapScore

    if homeResult > awayResult:
        changeHomeScore = int(homeScore)
        changeAwayScore = changeHomeScore - gapScore

        changeHomeHalfScore = halfList[0]
        changeAwayHalfScore = halfList[0] - halfGapScore
    else:
        changeAwayScore = int(awayScore)
        changeHomeScore = changeAwayScore - gapScore

        changeAwayHalfScore = halfList[1]
        changeHomeHalfScore = halfList[1] - halfGapScore
        
    print(homeName + ' : ' + awayName)
    print('기본 전반전 점수 -> ' + str(halfList[0]) + ' : ' + str(halfList[1]))
    print('스탯 점수 추가 전반전 -> ' + str(changeHomeHalfScore) + ' : ' + str(changeAwayHalfScore))
    print('기본 -> ' + str(int(homeScore)) + ' : ' + str(int(awayScore)))
    print('스탯 점수 추가 후 -> ' + str(changeHomeScore) + ' : ' + str(changeAwayScore))
    print('총 합 스탯 점수 -> ' + str(homeResult) + ' : ' + str(awayResult))
    print('팀 스탯 점수 -> ' + str(homeTeamScore) + ' : ' + str(awayTeamScore))
    print('플레이어 스탯 점수 -> ' + str(homePlayerScore) + ' : ' + str(awayPlayerScore))
    print('■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■')

if __name__ == '__main__':

    # ■■■■■■■■■■ 매일 업데이트 ■■■■■■■■■■
    
    # wkbl_Team_DataLoad(2023)
    # wkbl_Player_DataLoad(2023)
    # betman_WKBL_DataLoad(2023)

    # wkbl_Team_Analysis()
    # wkbl_Player_Analysis()

    # ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■

    # 배트맨 라인업
    betman_DailyLineup_Analysis('삼성생명','하나원큐')

