from requests_html import HTMLSession
import requests
import json
import urllib
import openpyxl
import pandas as pd
import time
from wordpress_xmlrpc import Client
from wordpress_xmlrpc import WordPressPost
from wordpress_xmlrpc.methods import posts
from openpyxl import load_workbook
from datetime import datetime
import os.path   
import random
import pprint
import datetime
from collections import Counter
import numpy as np

betman_mkdir_path = 'D:/pythonAppCode/ProtoAnalysis/Betman/'

# 베트맨(프로토사이트) 데이터 크롤링
def betman_DataLoad(year):

    fileName = betman_mkdir_path + '배트맨_' + str(year) + '년도_데이터.xlsx'

    print(fileName)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['No', '종류', '날짜', '리그 이름', '홈 네임', '원정 네임','승리 배당','무 배당','패배 배당','홈 핸디','무 핸디','원정 핸디','경기 결과','경기 결과 점수'])

    wb.save(fileName)  

    year = str(year)[2:]

    url = "https://www.betman.co.kr/buyPsblGame/gameInfoInq.do"

    # 210001 = 21년도 / 0001 -> 1주차 경기
    dataNumbering = int(year+'0001')

    row = 2
    for i in range(200):

        print('데이터 주차 : ' + str(dataNumbering))

        headers = {
            "Referer": f"https://www.betman.co.kr/main/mainPage/gamebuy/gameSlipIFR.do?gmId=G101&gmTs={str(dataNumbering)}&gameDivCd=C&isIFR=Y",            
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36",
            'accept': '*/*',
            'accept-encoding': 'gzip, deflate, br',
            'accept-language': 'ko',
            'Content-Type': 'application/json'
        }

        data = {
            "gmId" : "G101",
            "gmTs" : str(dataNumbering),
            "gameYear" : "",
            "_sbmInfo" : {"_sbmInfo" : {"debugMode": "false"}}
        }

        with HTMLSession() as s:
            result = s.post(url, headers=headers, data=json.dumps(data)).text
            json_result = json.loads(result)

            try:
                output = json_result['compSchedules']['datas']
            except:
                print(str(dataNumbering) + ' : 데이터가 없습니다.')
                break

            dataNumbering += 1

            for li in output:

                _type = li[0] # 스포츠 종류 (SC=축구/BK=농구/VL=배구)

                # 경기 날짜, 시간
                date = str(li[3])[:len(str(li[3]))-3] # 13자리에서 뒤에 3자리 빼고 변환(파이썬에서는 그렇게 해야 변환되서)
                _matchDate = datetime.datetime.fromtimestamp(int(date)).strftime('%Y-%m-%d %H:%M:%S')

                _leagueName = li[7] # 리그 이름
                _homeName = li[14] # 홈 네임
                _awayName = li[15] # 원정 네임
                _winRate = li[16] # 승리 배당률
                _drawRate = li[17] # 무승부 배당률
                _lossRate = li[18] # 패배 배당률

                _homeHandy = li[20] # 홈 핸디
                _drawHandy = li[21] # 무 핸디
                _awayHandy = li[22] # 원정 핸디

                _result = li[28] # 경기 결과 - 홈 0 / 무 1 / 원정 2

                _matchResult = li[33] # 경기 결과 점수

                # 경기 결과 점수가 없으면 넘기기
                if _matchResult == None: 
                    continue

                ## 엑셀 저장
                ws.cell(row, 1).value = row-1
                ws.cell(row, 2).value = _type
                ws.cell(row, 3).value = _matchDate
                ws.cell(row, 4).value = _leagueName
                ws.cell(row, 5).value = _homeName
                ws.cell(row, 6).value = _awayName
                ws.cell(row, 7).value = _winRate
                ws.cell(row, 8).value = _drawRate
                ws.cell(row, 9).value = _lossRate
                ws.cell(row, 10).value = _homeHandy
                ws.cell(row, 11).value = _drawHandy
                ws.cell(row, 12).value = _awayHandy
                ws.cell(row, 13).value = _result
                ws.cell(row, 14).value = _matchResult
                row += 1

    wb.save(fileName)

# 배당에 따른 경기 데이터 분석 통계
def betman_Match_Analysis():

    yearList = [2014,2015,2016,2017,2018,2019,2020,2021,2022,2023]

    winList = []
    lossList = []

    for year in yearList:

        loadfileName = betman_mkdir_path + '배트맨_' + str(year) + '년도_데이터.xlsx'

        wb = load_workbook(filename = loadfileName, data_only=True)
        ws = wb[wb.sheetnames[0]]

        print(loadfileName)

        for row in ws.rows:

            # 0='No', 1='종류', 2='날짜', 3='리그 이름', 4='홈 네임', 5='원정 네임', 6='승리 배당'
            # 7='무 배당', 8='패배 배당', 9='홈 핸디', 10='무 핸디', 11='원정 핸디', 12='경기 결과', 12='경기 결과 점수'

            _homeRate = str(row[6].value)
            _drawRate = str(row[7].value)
            _awayRate = str(row[8].value)

            # 0 = 홈팀 승, 1 = 무승부, 2 = 원정팀 승
            _result = str(row[12].value)

            if _result == '0': 
                winList.append(_homeRate)

                lossList.append(_drawRate)
                lossList.append(_awayRate)
            elif _result == '1': 
                winList.append(_drawRate)

                lossList.append(_homeRate)
                lossList.append(_awayRate)
            elif _result == '2': 
                winList.append(_awayRate)

                lossList.append(_homeRate)
                lossList.append(_drawRate)

    winList = Counter(winList)
    lossList = Counter(lossList)

    fileName = betman_mkdir_path + '배트맨_분석_확률_데이터.xlsx'
    print(fileName)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['배당률', '승리','패배','확률'])
    wb.save(fileName)  

    index = 2
    for key, value in winList.items():

        ws.cell(index, 1).value = key
        ws.cell(index, 2).value = value
        ws.cell(index, 3).value = lossList[key]
        ws.cell(index, 4).value = round((value / (value + lossList[key]))*100, 1)
        index += 1

    wb.save(fileName)  

    data=pd.read_excel(fileName) #원본 엑셀 파일 
    data=data.sort_values(by='확률', ascending=False) # 리뷰개수로 내림차순 정렬

    with pd.ExcelWriter(fileName) as writer:
        data.to_excel(writer,sheet_name="sheet0",index=False) # 그대로 저장

# 베트맨(프로토사이트) 주차별 데이터 크롤링
def betman_Today_Analysis(week):

    loadfileName = betman_mkdir_path + '배트맨_분석_확률_데이터.xlsx'

    wb = load_workbook(filename = loadfileName, data_only=True)
    ws = wb[wb.sheetnames[0]]

    url = "https://www.betman.co.kr/buyPsblGame/gameInfoInq.do"

    print('데이터 주차 : ' + str(week))

    headers = {
        "Referer": f"https://www.betman.co.kr/main/mainPage/gamebuy/gameSlipIFR.do?gmId=G101&gmTs={str(week)}&gameDivCd=C&isIFR=Y",            
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36",
        'accept': '*/*',
        'accept-encoding': 'gzip, deflate, br',
        'accept-language': 'ko',
        'Content-Type': 'application/json'
    }

    data = {
        "gmId" : "G101",
        "gmTs" : str(week),
        "gameYear" : "",
        "_sbmInfo" : {"_sbmInfo" : {"debugMode": "false"}}
    }

    with HTMLSession() as s:
        result = s.post(url, headers=headers, data=json.dumps(data)).text
        json_result = json.loads(result)

        try:
            output = json_result['compSchedules']['datas']
        except:
            print(str(week) + ' : 데이터가 없습니다.')
            return

        textCheck = ''

        percentageList = []

        getRateNumber = 1
        index = 1

        perNumber = 1

        for li in output:

            # 0='No', 1='종류', 2='날짜', 3='리그 이름', 4='홈 네임', 5='원정 네임', 6='승리 배당'
            # 7='무 배당', 8='패배 배당', 9='홈 핸디', 10='무 핸디', 11='원정 핸디', 12='경기 결과', 12='경기 결과 점수'

            _type = li[0] # 스포츠 종류 (SC=축구/BK=농구/VL=배구)

            # 경기 날짜, 시간
            date = str(li[3])[:len(str(li[3]))-3] # 13자리에서 뒤에 3자리 빼고 변환(파이썬에서는 그렇게 해야 변환되서)
            _matchDate = datetime.datetime.fromtimestamp(int(date)).strftime('%Y-%m-%d %H:%M:%S')

            _homeName = li[14] # 홈 네임

            _winRate = li[16] # 승리 배당률
            _drawRate = li[17] # 무승부 배당률
            _lossRate = li[18] # 패배 배당률

            _homeHandy = li[20] # 홈 핸디

            _matchResult = li[33] # 경기 결과 점수

            if _matchResult != None or _winRate == 0.0 : continue

            for row in ws.rows:
                if row[0].value == _winRate: 
                    homePercentage = row[3].value

                if row[0].value == _drawRate: 
                    drawPercentage = row[3].value

                if row[0].value == _lossRate: 
                    awayPercentage = row[3].value

            if _drawRate == 0.0: drawPercentage = 0.0

            if textCheck != _homeName:
                textCheck = _homeName

                if len(percentageList) != 0:
                    if index == 11:
                        print('■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■')
                        print('곱한 배당값 : ', getRateNumber)
                        # print('확률이 제일 높은 것을 곱한 값 : ' + str(perNumber))
                        print(f'배당에 대한 1000원 곱한 값 ★ {str(index-1)} : ', str(format(int(result), ',')) + '원')

                        getRateNumber = 1
                        perNumber = 1
                        index = 1

                    # 첫번째 큰 숫자 삭제
                    max1 = max(percentageList)
                    percentageList.remove(max1)

                    # 두번째 큰 숫자 삭제
                    max2 = max(percentageList)
                    percentageList.remove(max2)

                    # 세번째 큰 숫자 삭제
                    max3 = max(percentageList)
                    percentageList.remove(max3)

                    # 네번째 큰 숫자 삭제
                    max4 = max(percentageList)
                    percentageList.remove(max4)

                    resultNumber = max(percentageList)

                    for row in ws.rows:
                        if row[3].value == resultNumber: 
                            rateNumber = row[0].value
                            break

                    percentageList = []

                    getRateNumber = getRateNumber * rateNumber

                    result = getRateNumber * 1000

                    perNumber = perNumber * (resultNumber/100)
                    index += 1
                
                percentageList.append(homePercentage)
                if drawPercentage != 0.0: percentageList.append(drawPercentage)
                percentageList.append(awayPercentage)

                # print('■■■■■■■■■■■■■■■■■■■■■')
            else:
                percentageList.append(homePercentage)
                if drawPercentage != 0.0: percentageList.append(drawPercentage)
                percentageList.append(awayPercentage)

            # print(_type + '/' + _matchDate + '/' + _homeName + '/핸디:'+ str(_homeHandy) + ' / ◀  홈승:' + str(_winRate) + ' 확률=> '+str(homePercentage) + '% ▶ / ◀  무승부:' + str(_drawRate) + ' 확률=> '+str(drawPercentage) + '% ▶ / ◀  원정승' + str(_lossRate) + ' 확률=> ' + str(awayPercentage)+'% ▶')

if __name__ == '__main__':
    # betman_DataLoad(2023)
    # betman_Match_Analysis()
    betman_Today_Analysis(230019)

